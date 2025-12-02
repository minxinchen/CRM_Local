#!/usr/bin/env python3
"""
建立 CRM 系統的 Excel 範本檔案
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def create_tasks_template():
    """建立任務資料範本"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    
    # 設定標題列
    headers = ["ID", "標題", "描述", "狀態", "到期日", "建立時間", "完成時間"]
    ws.append(headers)
    
    # 格式化標題列
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    
    # 新增範例資料
    ws.append([
        1,
        "找總經理討論 SN25038",
        "需要確認機台規格與報價",
        "critical",
        "",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ""
    ])
    
    wb.save("../data/tasks.xlsx")
    print("✓ tasks.xlsx 已建立")

def create_customers_template():
    """建立客戶資料範本"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # 設定標題列
    headers = ["ID", "客戶名稱", "Email", "電話", "公司", "網域", "狀態", "備註", "背景資訊", "建立時間"]
    ws.append(headers)
    
    # 格式化標題列
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 40
    ws.column_dimensions['J'].width = 18
    
    # 新增範例資料
    ws.append([
        1,
        "Test Sender",
        "test@example.com",
        "",
        "",
        "example.com",
        "new_lead",
        "",
        "",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])
    
    wb.save("../data/customers.xlsx")
    print("✓ customers.xlsx 已建立")

def create_cases_template():
    """建立案件資料範本"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"
    
    # 設定標題列
    headers = ["ID", "客戶ID", "標題", "描述", "狀態", "優先級", "建立時間", "更新時間"]
    ws.append(headers)
    
    # 格式化標題列
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    wb.save("../data/cases.xlsx")
    print("✓ cases.xlsx 已建立")

def create_emails_template():
    """建立郵件資料範本"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"
    
    # 設定標題列
    headers = ["ID", "MessageID", "主旨", "寄件者Email", "寄件者名稱", "收件者Email", "內容", "收件時間", "客戶ID", "狀態"]
    ws.append(headers)
    
    # 格式化標題列
    header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    
    # 新增範例資料
    ws.append([
        1,
        "sample-message-id-001",
        "詢價：CNC 加工機",
        "customer@example.com",
        "張先生",
        "sales@mycompany.com",
        "您好，我們想詢問貴公司的 CNC 加工機價格與規格...",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        1,
        "unprocessed"
    ])
    
    wb.save("../data/emails.xlsx")
    print("✓ emails.xlsx 已建立")

def create_interactions_template():
    """建立溝通紀錄範本"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Interactions"
    
    # 設定標題列
    headers = ["ID", "案件ID", "類型", "內容", "溝通時間"]
    ws.append(headers)
    
    # 格式化標題列
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 設定欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 18
    
    wb.save("../data/interactions.xlsx")
    print("✓ interactions.xlsx 已建立")

if __name__ == "__main__":
    os.chdir(os.path.dirname(__file__))
    
    print("正在建立 Excel 範本檔案...")
    create_tasks_template()
    create_customers_template()
    create_cases_template()
    create_emails_template()
    create_interactions_template()
    print("\n✅ 所有 Excel 範本檔案已建立完成！")
